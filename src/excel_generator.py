"""Excel template generator for supply and demand planning."""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
import os


class ExcelGenerator:
    """Generates Excel templates for supply and demand planning."""
    
    def __init__(self, output_dir: str = "templates"):
        """Initialize Excel generator."""
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    def create_planning_template(self, filename: str = "supply_demand_plan.xlsx"):
        """Create a blank planning template."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Planning"
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['Date', 'Current Stock', 'Historical Avg Daily Demand', 
                   'Planned Sales', 'Forecasted Demand', 'Safety Stock', 
                   'Reorder Point', 'Projected Stock', 'Purchase Needed', 'Suggested Purchase']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Add sample rows (12 months)
        for row in range(2, 14):
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                
                # Format numbers
                if col in [2, 3, 4, 5, 6, 7, 8, 10]:
                    cell.number_format = '0'
                elif col == 1:
                    cell.number_format = 'yyyy-mm-dd'
                
                # Add formulas for calculations (examples)
                if col == 7 and row == 2:  # Reorder Point
                    cell.value = f"=C{row}+F{row}"
                elif col == 8 and row == 2:  # Projected Stock
                    cell.value = f"=B{row}-E{row}"
                elif col == 9 and row == 2:  # Purchase Needed
                    cell.value = f"=IF(H{row}<G{row},TRUE,FALSE)"
                elif col == 10 and row == 2:  # Suggested Purchase
                    cell.value = f"=IF(I{row},G{row}-H{row},0)"
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Add "Input Data" sheet
        ws_input = wb.create_sheet("Input Data")
        input_headers = ['Description', 'Value']
        
        for col, header in enumerate(input_headers, 1):
            cell = ws_input.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        input_data = [
            ['Current Stock Level', ''],
            ['Safety Stock Level', ''],
            ['Lead Time (days)', ''],
            ['Average Daily Demand (baseline)', ''],
        ]
        
        for idx, (desc, val) in enumerate(input_data, 2):
            ws_input.cell(row=idx, column=1, value=desc).border = border
            ws_input.cell(row=idx, column=2, value=val).border = border
            ws_input.cell(row=idx, column=2).number_format = '0'
        
        ws_input.column_dimensions['A'].width = 30
        ws_input.column_dimensions['B'].width = 15
        
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        print(f"Template created: {filepath}")
        return filepath
    
    def _auto_adjust_columns(self, worksheet):
        """Auto-adjust worksheet column widths."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _style_header_row(self, worksheet, color: str = "70AD47"):
        """Apply standard styling to the first row of a worksheet."""
        header_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col in worksheet.iter_cols(min_row=1, max_row=1):
            for cell in col:
                cell.fill = header_fill
                cell.font = header_font

    def create_forecast_report(
        self,
        forecast_df: pd.DataFrame,
        filename: str = "forecast_report.xlsx",
        action_list_df: pd.DataFrame | None = None,
        stock_outlook_df: pd.DataFrame | None = None,
        dashboard_title: str = "Planner Dashboard",
    ):
        """Create a multi-sheet planning report with forecast, actions, and dashboard."""
        filepath = os.path.join(self.output_dir, filename)

        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            forecast_df.to_excel(writer, sheet_name='Forecast', index=False)

            workbook = writer.book
            forecast_ws = writer.sheets['Forecast']
            self._style_header_row(forecast_ws, color="70AD47")
            self._auto_adjust_columns(forecast_ws)

            if action_list_df is not None:
                action_list_df.to_excel(writer, sheet_name='Action List', index=False)
                action_ws = writer.sheets['Action List']
                self._style_header_row(action_ws, color="2F5597")
                self._auto_adjust_columns(action_ws)

                if stock_outlook_df is not None:
                    stock_outlook_df.to_excel(writer, sheet_name='Stock Outlook', index=False)
                    stock_ws = writer.sheets['Stock Outlook']
                    self._style_header_row(stock_ws, color="9C6500")
                    self._auto_adjust_columns(stock_ws)

                dashboard_ws = workbook.create_sheet("Dashboard")
                dashboard_ws["A1"] = dashboard_title
                dashboard_ws["A1"].font = Font(size=16, bold=True, color="1F4E78")

                kpi_items = [
                    ("Total Forecast Rows", len(forecast_df)),
                    ("SKUs in Plan", int(action_list_df["sku"].nunique()) if "sku" in action_list_df.columns else 1),
                    ("SKUs To Buy Now", int(action_list_df["buy_now"].sum()) if "buy_now" in action_list_df.columns else 0),
                    ("High Risk SKUs", int((action_list_df["risk_level"] == "high").sum()) if "risk_level" in action_list_df.columns else 0),
                    ("Total Recommended Qty", float(action_list_df["total_recommended_qty_horizon"].sum()) if "total_recommended_qty_horizon" in action_list_df.columns else 0.0),
                ]

                dashboard_ws["A3"] = "KPI"
                dashboard_ws["B3"] = "Value"
                dashboard_ws["A3"].font = Font(bold=True, color="FFFFFF")
                dashboard_ws["B3"].font = Font(bold=True, color="FFFFFF")
                dashboard_ws["A3"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                dashboard_ws["B3"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

                for idx, (kpi, value) in enumerate(kpi_items, start=4):
                    dashboard_ws[f"A{idx}"] = kpi
                    dashboard_ws[f"B{idx}"] = value

                dashboard_ws["A10"] = "Top Urgent SKUs"
                dashboard_ws["A10"].font = Font(size=12, bold=True, color="1F4E78")
                urgent_cols = ["sku", "buy_now", "order_by_date", "expected_stockout_date", "risk_level", "initial_recommended_qty", "primary_reason_code"]
                urgent_cols = [col for col in urgent_cols if col in action_list_df.columns]

                urgent_df = action_list_df.copy()
                if "risk_level" in urgent_df.columns:
                    risk_rank = {"high": 0, "medium": 1, "low": 2}
                    urgent_df["_risk_rank"] = urgent_df["risk_level"].map(risk_rank).fillna(9)
                    urgent_df = urgent_df.sort_values(
                        by=["_risk_rank", "buy_now", "order_by_date", "initial_recommended_qty"],
                        ascending=[True, False, True, False],
                    ).drop(columns=["_risk_rank"])
                urgent_df = urgent_df.head(20)

                start_row = 11
                for col_idx, col_name in enumerate(urgent_cols, start=1):
                    cell = dashboard_ws.cell(row=start_row, column=col_idx, value=col_name)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="8064A2", end_color="8064A2", fill_type="solid")

                for row_offset, (_, row_data) in enumerate(urgent_df[urgent_cols].iterrows(), start=1):
                    for col_idx, col_name in enumerate(urgent_cols, start=1):
                        dashboard_ws.cell(row=start_row + row_offset, column=col_idx, value=row_data[col_name])

                self._auto_adjust_columns(dashboard_ws)

        print(f"Report created: {filepath}")
        return filepath
