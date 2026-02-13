"""Append planner output sheets and dashboard to an existing FS workbook."""

from pathlib import Path
import shutil

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill


DEFAULT_INPUT = Path(r"C:\Users\elmer\Downloads\FS 2025_2026 (obv Odoo Files).xlsx")
DEFAULT_OUTPUT = Path("templates/FS 2025_2026 (obv Odoo Files) - enhanced.xlsx")


def _build_kpi_frame(results: dict) -> pd.DataFrame:
    purchase_plan = results["purchase_plan"]
    action_list = results["action_list"]
    stock_outlook = results["stock_outlook"]

    total_buy_qty = float(
        purchase_plan.loc[purchase_plan["purchase_needed"], "suggested_purchase"].sum()
    )
    stockout_periods = int(purchase_plan["stockout_risk"].sum())
    late_po_periods = int(purchase_plan["late_po_risk"].sum())
    skus = int(purchase_plan["sku"].nunique()) if "sku" in purchase_plan.columns else 1
    buy_now = int(action_list["buy_now"].sum()) if "buy_now" in action_list.columns else 0
    high_risk_skus = (
        int((action_list["risk_level"] == "high").sum()) if "risk_level" in action_list.columns else 0
    )
    medium_risk_skus = (
        int((action_list["risk_level"] == "medium").sum()) if "risk_level" in action_list.columns else 0
    )
    low_risk_skus = (
        int((action_list["risk_level"] == "low").sum()) if "risk_level" in action_list.columns else 0
    )

    rows = [
        {"KPI": "SKUs in plan", "Value": skus},
        {"KPI": "Action list rows", "Value": len(action_list)},
        {"KPI": "Stock outlook rows", "Value": len(stock_outlook)},
        {"KPI": "Buy now SKUs", "Value": buy_now},
        {"KPI": "High risk SKUs", "Value": high_risk_skus},
        {"KPI": "Medium risk SKUs", "Value": medium_risk_skus},
        {"KPI": "Low risk SKUs", "Value": low_risk_skus},
        {"KPI": "Stockout risk periods", "Value": stockout_periods},
        {"KPI": "Late PO risk periods", "Value": late_po_periods},
        {"KPI": "Total suggested purchase qty", "Value": round(total_buy_qty, 2)},
    ]
    return pd.DataFrame(rows)


def _write_planner_sheets(output_path: Path, results: dict):
    action_list = results["action_list"].copy()
    stock_outlook = results["stock_outlook"].copy()
    purchase_plan = results["purchase_plan"].copy()
    kpi_df = _build_kpi_frame(results)

    urgent = action_list.copy()
    if "risk_level" in urgent.columns:
        rank = {"high": 0, "medium": 1, "low": 2}
        urgent["_risk_rank"] = urgent["risk_level"].map(rank).fillna(9)
        urgent = urgent.sort_values(
            by=["_risk_rank", "buy_now", "order_by_date", "total_recommended_qty_horizon"],
            ascending=[True, False, True, False],
        ).drop(columns=["_risk_rank"])
    urgent = urgent.head(50)

    with pd.ExcelWriter(
        output_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
    ) as writer:
        kpi_df.to_excel(writer, sheet_name="Planner_KPI", index=False)
        urgent.to_excel(writer, sheet_name="Planner_Urgent", index=False)
        action_list.to_excel(writer, sheet_name="Planner_Action_List", index=False)
        stock_outlook.to_excel(writer, sheet_name="Planner_Stock_Outlook", index=False)
        purchase_plan.to_excel(writer, sheet_name="Planner_Purchase_Plan", index=False)


def _build_planner_dashboard(output_path: Path):
    wb = load_workbook(output_path)
    if "Planner_Dashboard" in wb.sheetnames:
        del wb["Planner_Dashboard"]
    ws = wb.create_sheet("Planner_Dashboard")

    title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    white_bold = Font(color="FFFFFF", bold=True)

    ws["A1"] = "Planner Dashboard (auto-linked to planner sheets)"
    ws["A1"].font = Font(size=14, bold=True, color="1F4E78")

    ws["A3"] = "KPI"
    ws["B3"] = "Value"
    ws["A3"].fill = header_fill
    ws["B3"].fill = header_fill
    ws["A3"].font = white_bold
    ws["B3"].font = white_bold

    kpi_labels = [
        "SKUs in plan",
        "Buy now SKUs",
        "High risk SKUs",
        "Medium risk SKUs",
        "Low risk SKUs",
        "Stockout risk periods",
        "Late PO risk periods",
        "Total suggested purchase qty",
    ]
    for i, label in enumerate(kpi_labels, start=4):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = f'=IFERROR(INDEX(Planner_KPI!$B:$B, MATCH("{label}", Planner_KPI!$A:$A, 0)), "")'

    ws["D3"] = "Top Urgent SKUs"
    ws["D3"].font = Font(size=12, bold=True, color="1F4E78")
    urgent_headers = ["sku", "ART_NAME", "SIZE_CODE US", "risk_level", "buy_now", "order_by_date", "initial_recommended_qty"]
    for col_idx, head in enumerate(urgent_headers, start=4):
        cell = ws.cell(row=4, column=col_idx, value=head)
        cell.fill = title_fill
        cell.font = white_bold

    for row in range(5, 55):
        source_row = row - 3
        ws.cell(row=row, column=4, value=f"=Planner_Urgent!A{source_row}")
        ws.cell(row=row, column=5, value=f"=Planner_Urgent!J{source_row}")
        ws.cell(row=row, column=6, value=f"=Planner_Urgent!O{source_row}")
        ws.cell(row=row, column=7, value=f"=Planner_Urgent!G{source_row}")
        ws.cell(row=row, column=8, value=f"=Planner_Urgent!B{source_row}")
        ws.cell(row=row, column=9, value=f"=Planner_Urgent!C{source_row}")
        ws.cell(row=row, column=10, value=f"=Planner_Urgent!D{source_row}")

    ws["A14"] = "Risk Level"
    ws["B14"] = "SKUs"
    ws["A14"].fill = header_fill
    ws["B14"].fill = header_fill
    ws["A14"].font = white_bold
    ws["B14"].font = white_bold
    ws["A15"] = "High"
    ws["B15"] = '=IFERROR(INDEX(Planner_KPI!$B:$B, MATCH("High risk SKUs", Planner_KPI!$A:$A, 0)), 0)'
    ws["A16"] = "Medium"
    ws["B16"] = '=IFERROR(INDEX(Planner_KPI!$B:$B, MATCH("Medium risk SKUs", Planner_KPI!$A:$A, 0)), 0)'
    ws["A17"] = "Low"
    ws["B17"] = '=IFERROR(INDEX(Planner_KPI!$B:$B, MATCH("Low risk SKUs", Planner_KPI!$A:$A, 0)), 0)'

    chart = BarChart()
    chart.title = "SKU Risk Distribution"
    data = Reference(ws, min_col=2, min_row=14, max_row=17)
    categories = Reference(ws, min_col=1, min_row=15, max_row=17)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 6
    chart.width = 10
    ws.add_chart(chart, "A19")

    widths = {
        "A": 28,
        "B": 14,
        "D": 12,
        "E": 26,
        "F": 12,
        "G": 12,
        "H": 10,
        "I": 14,
        "J": 20,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(output_path)


def enhance_fs_workbook(results: dict, input_path: Path = DEFAULT_INPUT, output_path: Path = DEFAULT_OUTPUT):
    """Create an enhanced copy of the FS workbook with planning sheets and dashboard."""
    if not input_path.exists():
        raise FileNotFoundError(f"Input workbook not found: {input_path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(input_path, output_path)
    _write_planner_sheets(output_path, results)
    _build_planner_dashboard(output_path)
    return output_path


if __name__ == "__main__":
    from main import main as run_planner

    planner_results = run_planner()
    enhanced_path = enhance_fs_workbook(planner_results)
    print(f"Enhanced workbook created: {enhanced_path}")
